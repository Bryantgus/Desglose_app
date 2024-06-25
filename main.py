import ttkbootstrap as tb
from tkinter import ttk
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from ttkbootstrap import *
from fractions import Fraction
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font
import shutil
import win32com.client as win32

# Styles
#morph
THEME_WINDOW = "morph"
TITLE_WINDOW = "Desglose P65"
THEME_LABEL_TITLE = "dark"
FRAME_DESGLOSE = "primary"
ENTRY_STYLE = "info"
MAIN_FRAME_STYLE = "dark"
WIDTH_LABELS_FRAME_DESGLOSE = 7
#
FONT_COLOR_LETTERS = "#01204E"
FONT_LETTERS = ("Arial", 11, "bold")


class App(tb.Window):
    def __init__(self):
        super().__init__(themename=THEME_WINDOW)
        self.fila_cantidad = 2
        self.title(TITLE_WINDOW)
        self.geometry("750x700+300+0")

        # Crear el frame dentro del canvas
        self.main_frame = Frame(self, bootstyle="info", padding=3 )
        self.main_frame.grid(row=0, column=0, columnspan=3)

        # Diccionarios que almacena los valores que cambiaran de cada frame
        self.ancho_entry = {}
        self.alto_entry = {}
        self.num_entry = {}
        self.labels_results = {}
        # Dicionarios que almacenan los datos que se introduciran en los frames
        self.ancho_values = {}
        self.alto_values = {}
        self.ryc = {}
        self.r = {}
        self.l = {}
        self.j = {}
        self.can = {}
        self.cal = {}
        self.results = {}
        self.starting_app()
        self.type_desglose = 1
        self.actual_label = 1

    def starting_app(self):
        self.menu()
        self.text_title()
        self.painting_frame()
        self.buttons()

    def menu(self):
        menubar = Menu(self)
        self.config(menu=menubar)
        export_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Archivo", menu=export_menu)
        export_menu.add_command(label="Exportar", command=self.export_to_excel)
        export_menu.add_command(label="Imprimir", command=self.print_sheets)

    def print_sheets(self):
        # Ruta del archivo Excel
        ruta_archivo = "C:/Users/EJ/PycharmProjects/Desglose/Formato desglose.xlsx"

        # Inicializar Excel a través de COM
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False  # Hacer Excel no visible

        # Abrir el archivo
        libro = excel.Workbooks.Open(ruta_archivo)

        # Lista de nombres de hojas
        nombres_hojas = [hoja.Name for hoja in libro.Sheets]
        print("Hojas en el archivo:", nombres_hojas)
        uno_tres = libro.Sheets['1-3']
        uno_seis = libro.Sheets['1-6']
        uno_nueve = libro.Sheets['1-9']
        uno_doce = libro.Sheets['1-12']
        trece_quince = libro.Sheets['13-15']
        trece_diezocho = libro.Sheets['13-18']
        trece_ventiuno = libro.Sheets['13-21']
        trece_venticuatro = libro.Sheets['13-24']
        venticinco_siete = libro.Sheets['1-12']
        veinticindo_cero = libro.Sheets['1-12']
        venticinco_tres = libro.Sheets['1-12']
        venticinco_seis = libro.Sheets['1-12']



        # else:
        #     uno_seis.PrintOut()
        libro.Close(SaveChanges=False)
        excel.Quit()

    def text_title(self):
        opciones = ["P 65, 2 Vías", "P 65, 3 Vías", "Tradicional, 2 Vías", "Tradicional, 3 Vías"]
        options = ttk.Combobox(self.main_frame, values=opciones, bootstyle="dark")
        options.configure(cursor='hand2', state='readonly')
        options.set(opciones[0])
        options.grid(row=0, column=0, pady=10)
        options.bind("<<ComboboxSelected>>", self.on_combobox_change)

    def on_combobox_change(self, event):
        selected_option = event.widget.get()
        if selected_option == "P 65, 2 Vías":
            self.type_desglose = 1
        elif selected_option == "P 65, 3 Vías":
            self.type_desglose = 2
        elif selected_option == "Tradicional, 2 Vías":
            self.type_desglose = 3
        elif selected_option == "Tradicional, 3 Vías":
            self.type_desglose = 4
        self.calculate_values()
        print(self.type_desglose)

    def painting_frame(self):
        num = 1
        for row in range(1, 3):
            for column in range(0, 4):
                self.frame_desglose(num, row, column)
                num += 1

    def frame_desglose(self, num, row, column):
        frame_desglose = Frame(self.main_frame, bootstyle=FRAME_DESGLOSE, padding=10)
        frame_desglose.grid(row=row, column=column, padx=5, pady=5)

        numero_label = tb.Label(frame_desglose, text="Nº", anchor="center",
                                width=WIDTH_LABELS_FRAME_DESGLOSE,
                                font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        numero_label.grid(row=0, column=0)

        numero_desglose = tb.Entry(frame_desglose, justify="center", font=FONT_LETTERS, width=3,
                                   bootstyle=ENTRY_STYLE, foreground=FONT_COLOR_LETTERS)
        numero_desglose.grid(row=0, column=1)
        self.num_entry[f"entry_{num}"] = numero_desglose

        numero_desglose.insert(0, num)

        ancho_label = tb.Label(frame_desglose, text="Ancho", anchor="center",
                               width=WIDTH_LABELS_FRAME_DESGLOSE,
                               font=FONT_LETTERS,
                               foreground=FONT_COLOR_LETTERS)
        ancho_label.grid(row=1, column=0, pady=2)

        alto_label = tb.Label(frame_desglose, bootstyle="secondary", text="Alto", anchor="center",
                              width=WIDTH_LABELS_FRAME_DESGLOSE,
                              font=FONT_LETTERS,
                              foreground=FONT_COLOR_LETTERS)
        alto_label.grid(row=1, column=1, pady=2)

        # Crear entries
        ancho_entry = tb.Entry(frame_desglose, justify="center", font=FONT_LETTERS, width=7,
                               bootstyle=ENTRY_STYLE,
                               foreground=FONT_COLOR_LETTERS)
        ancho_entry.grid(row=2, column=0, pady=5)
        self.ancho_entry[f'ancho_{num}'] = ancho_entry

        alto_entry = tb.Entry(frame_desglose, justify="center", font=FONT_LETTERS, width=7, bootstyle=ENTRY_STYLE,
                              foreground=FONT_COLOR_LETTERS)
        alto_entry.grid(row=2, column=1, pady=5)
        self.alto_entry[f'alto_{num}'] = alto_entry

        # Riel y Cabezal
        riel_cabezal = tb.Label(frame_desglose, text="RC", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE,
                                padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS, bootstyle="secondary")
        riel_cabezal.grid(row=3, column=0, padx=5, pady=1)
        # Ruleta
        ruleta = tb.Label(frame_desglose, text="Ruleta", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE,
                          padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        ruleta.grid(row=4, column=0, padx=5, pady=1)

        # Lateral
        lateral = tb.Label(frame_desglose, text="Lateral", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE,
                           padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        lateral.grid(row=5, column=0, padx=5, pady=1)

        # Jamba
        jamba = tb.Label(frame_desglose, text="Jamba", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE,
                         padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        jamba.grid(row=6, column=0, padx=5, pady=1)

        # Cristal Ancho
        cristal_ancho = tb.Label(frame_desglose, text="C.AN", anchor="center",
                                 width=WIDTH_LABELS_FRAME_DESGLOSE,
                                 padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        cristal_ancho.grid(row=7, column=0, padx=5, pady=1)

        # Cristal Largo
        cristal_alto = tb.Label(frame_desglose, text="C.AL", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE,
                                padding=2, font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        cristal_alto.grid(row=8, column=0, padx=5, pady=1)

        self.labels_results[f'desglose_{num}'] = {}
        # Riel y cabezal
        rc = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                      font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        rc.grid(row=3, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Riel y Cabezal'] = rc

        # Ruleta
        r = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                     font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        r.grid(row=4, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Ruleta'] = r

        # Lateral
        l = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                     font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        l.grid(row=5, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Lateral'] = l

        # Jamba
        j = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                     font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        j.grid(row=6, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Jamba'] = j

        # Cristal Ancho
        can = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                       font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        can.grid(row=7, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Cristal Ancho'] = can

        # Cristal Alto
        cal = tb.Label(frame_desglose, text="", anchor="center", width=WIDTH_LABELS_FRAME_DESGLOSE, padding=2,
                       font=FONT_LETTERS, foreground=FONT_COLOR_LETTERS)
        cal.grid(row=8, column=1, padx=5, pady=1)
        self.labels_results[f'desglose_{num}']['Cristal Alto'] = cal

    def buttons(self):

        calculate_button = tb.Button(self, text="Calcular", command=self.calculate_values,
                                     bootstyle="dark")
        calculate_button.grid(row=1, column=1)
        next_page = tb.Button(self, text=">", command=lambda: self.pagination(">"), bootstyle="dark")
        next_page.grid(row=1, column=2)
        back_page = tb.Button(self, text="<", command=lambda: self.pagination("<"), bootstyle="dark")
        back_page.grid(row=1, column=0)

    def saving_entrys(self):
        # Guardando los valores de los entrys alto-ancho
        actual_label = self.actual_label
        for key, value in self.ancho_entry.items():
            self.ancho_values[f'ancho_{actual_label}'] = self.ancho_entry[key].get()
            actual_label += 1
        actual_label = self.actual_label
        for key, value in self.alto_entry.items():
            self.alto_values[f'alto_{actual_label}'] = self.alto_entry[key].get()
            actual_label += 1

    def calculate_values(self):
        self.saving_entrys()
        for i in range(self.actual_label, self.actual_label + 8):
            clave_ancho = f"ancho_{i}"
            clave_alto = f"alto_{i}"

            # Obtener los valores correspondientes
            valor_ancho = str(self.ancho_values[clave_ancho])
            valor_alto = str(self.alto_values[clave_alto])
            self.mixto_math(valor_ancho, valor_alto, i, self.type_desglose)

        self.update_labels()

    def mixto_math(self, ancho, alto, num, type_desglose):
        # Verificar si ancho o alto son nulos, vacíos o 0
        resto_rc = 0
        resto_r = 0
        resto_l = 0
        resto_j = 0
        resto_can = 0
        resto_cal = 0
        two_or_three = 0
        if not ancho or ancho.strip() == "0" or not alto or alto.strip() == "0":
            self.results[f'desglose_{num}'] = {
                "Riel y Cabezal": "",
                "Ruleta": "",
                "Lateral": "",
                "Jamba": "",
                "Cristal Ancho": "",
                "Cristal Alto": ""
            }
            self.ryc = {f"Riel y Cabezal {num}": ""}
            self.r = {f"Ruleta {num}": ""}
            self.l = {f"Lateral {num}": ""}
            self.j = {f"Jamba {num}": ""}
            self.can = {f"Cristal Ancho {num}": ""}
            self.cal = {f"Cristal Alto {num}": ""}
            return
        if type_desglose == 1:
            # P65 2 vias
            resto_rc = "1 3/8"
            resto_r = "1 1/8"
            resto_l = "1/8"
            resto_j = "2 1/8"
            resto_can = "6 1/2"
            resto_cal = 5
            two_or_three = 2
        # P65 3 vias
        if type_desglose == 2:
            resto_rc = "1 3/8"
            resto_r = "3/8"
            resto_l = "1/8"
            resto_j = "2 1/8"
            resto_can = "7 3/8"
            resto_cal = 5
            two_or_three = 3
        # Tradicional 2 vias
        if type_desglose == 3:
            resto_rc = "1/8"
            resto_r = "1/2"
            resto_l = "1/2"
            resto_j = "1"
            resto_can = "4"
            resto_cal = 4
            two_or_three = 2
        # Tradicional 3 vias
        if type_desglose == 4:
            resto_rc = "1/8"
            resto_r = "0"
            resto_l = "1/2"
            resto_j = "1"
            resto_can = "6"
            resto_cal = 4
            two_or_three = 3

        # La f al final de la variable significa fraction
        ancho_f = sum(Fraction(s) for s in ancho.split())
        alto_f = sum(Fraction(s) for s in alto.split())

        # Riel y Cabezal
        resto_rc_f = sum(Fraction(s) for s in resto_rc.split())
        rc = ancho_f - resto_rc_f
        rc = self.decimal_to_fraction_inches(rc)

        # Ruleta
        r = 0
        print(r)
        resto_r_f = sum(Fraction(s) for s in resto_r.split())

        if type_desglose == 2:
            r = (ancho_f + resto_r_f) / two_or_three
        elif type_desglose == 3:
            r = ancho_f / two_or_three
        else:
            r = (ancho_f - resto_r_f) / two_or_three
        r = self.decimal_to_fraction_inches(r)

        # Lateral
        resto_l_f = sum(Fraction(s) for s in resto_l.split())
        l = alto_f - resto_l_f
        l = self.decimal_to_fraction_inches(l)

        # Jamba
        resto_j_f = sum(Fraction(s) for s in resto_j.split())
        j = alto_f - resto_j_f
        j = self.decimal_to_fraction_inches(j)

        # Cristal ancho
        resto_can_f = sum(Fraction(s) for s in resto_can.split())
        can = (ancho_f - resto_can_f) / two_or_three
        can = self.decimal_to_fraction_inches(can)

        # Cristal alto
        cal = alto_f - resto_cal
        cal = self.decimal_to_fraction_inches(cal)

        self.results[f'desglose_{num}'] = {
            "Riel y Cabezal": rc,
            "Ruleta": r,
            "Lateral": l,
            "Jamba": j,
            "Cristal Ancho": can,
            "Cristal Alto": cal
        }
        self.ryc = {f"Riel y Cabezal {num}": rc}
        self.r = {f"Ruleta {num}": r}
        self.l = {f"Lateral {num}": l}
        self.j = {f"Jamba {num}": j}
        self.can = {f"Cristal Ancho {num}": can}
        self.cal = {f"Cristal Alto {num}": cal}

    def update_labels(self):
        num = 1  # Para indexar los resultados
        for i in range(self.actual_label, self.actual_label + 8):
            desglose_key = f'desglose_{i}'

            if desglose_key in self.results:
                # Obtener los valores de los resultados
                riel_cabezal_value = self.results[desglose_key].get("Riel y Cabezal", "")
                ruleta_value = self.results[desglose_key].get("Ruleta", "")
                lateral_value = self.results[desglose_key].get("Lateral", "")
                jamba_value = self.results[desglose_key].get("Jamba", "")
                cristal_ancho_value = self.results[desglose_key].get("Cristal Ancho", "")
                cristal_alto_value = self.results[desglose_key].get("Cristal Alto", "")

                # Actualizar los Labels en la interfaz
                self.labels_results[f'desglose_{num}']["Riel y Cabezal"].config(text=riel_cabezal_value)
                self.labels_results[f'desglose_{num}']["Ruleta"].config(text=ruleta_value)
                self.labels_results[f'desglose_{num}'][("Late"
                                                        "ral")].config(text=lateral_value)
                self.labels_results[f'desglose_{num}']["Jamba"].config(text=jamba_value)
                self.labels_results[f'desglose_{num}']["Cristal Ancho"].config(text=cristal_ancho_value)
                self.labels_results[f'desglose_{num}']["Cristal Alto"].config(text=cristal_alto_value)

            num += 1

    @staticmethod
    def decimal_to_fraction_inches(value):
        # Separar la parte entera y la parte decimal
        whole_part = int(value)
        decimal_part = value - whole_part

        # Redondear la parte decimal a la fracción más cercana de 1/16
        fraction_part = round(decimal_part * 16) / 16
        fraction_part = Fraction(fraction_part).limit_denominator(16)

        # Combinar la parte entera con la fracción
        if fraction_part.numerator == 0:
            return f"{whole_part}"
        elif whole_part == 0:
            return f"{fraction_part}"
        else:
            return f"{whole_part} {fraction_part}"

    def pagination(self, next_or_previous):
        self.saving_entrys()
        self.calculate_values()
        # Guardando datos de los entrys ancho y alto
        if next_or_previous == "<":
            if self.actual_label != 1:
                self.actual_label -= 8
        elif next_or_previous == ">":
            self.actual_label += 8
        else:
            pass
        num = self.actual_label
        # Actualizando el número de desgloses
        for a in range(1, 9):
            self.num_entry[f'entry_{a}'].delete(0, tk.END)
            self.num_entry[f'entry_{a}'].insert(0, num)

            # Verificar si es la primera vez que se visita este rango de labels
            if f'ancho_{num}' not in self.ancho_values:
                # Primera vez: asignar valores nulos
                self.ancho_entry[f'ancho_{a}'].delete(0, tk.END)
                self.ancho_entry[f'ancho_{a}'].insert(0, "")
                self.alto_entry[f'alto_{a}'].delete(0, tk.END)
                self.alto_entry[f'alto_{a}'].insert(0, "")
            else:
                # Ya visitado: cargar los valores almacenados
                self.ancho_entry[f'ancho_{a}'].delete(0, tk.END)
                self.ancho_entry[f'ancho_{a}'].insert(0, self.ancho_values[f'ancho_{num}'])
                self.alto_entry[f'alto_{a}'].delete(0, tk.END)
                self.alto_entry[f'alto_{a}'].insert(0, self.alto_values[f'alto_{num}'])
            num += 1
        self.calculate_values()
        self.update_labels()

    @staticmethod
    def copy_excel_file(src, dest):
        shutil.copyfile(src, dest)

    def export_to_excel(self):
        archivo_excel = 'C:/Users/EJ/PycharmProjects/Desglose/Formato desglose.xlsx'
        nuevo_excel = 'C:/Users/EJ/PycharmProjects/Desglose/Formato copia.xlsx'
        self.copy_excel_file(archivo_excel, nuevo_excel)

        libro = load_workbook(nuevo_excel)
        hoja = libro['datos']

        fila = 1
        columna = 1

        for key, value in self.ancho_values.items():
            hoja.cell(row=fila, column=columna, value=key)
            hoja.cell(row=fila + 1, column=columna, value=value)
            fila += 2

        fila = 1
        columna = 2

        for key, value in self.alto_values.items():
            hoja.cell(row=fila, column=columna, value=key)
            hoja.cell(row=fila + 1, column=columna, value=value)
            fila += 2

        fila = 1
        for key, value in self.results.items():
            columna = 3
            for sub_key, sub_value in value.items():
                hoja.cell(row=fila, column=columna, value=sub_key)
                hoja.cell(row=fila + 1, column=columna, value=sub_value)
                columna += 1
            fila += 2

        libro.save(nuevo_excel)
        print("Archivo guardado  exitosamente.")


if __name__ == "__main__":
    app = App()
    app.mainloop()
